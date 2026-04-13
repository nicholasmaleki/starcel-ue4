"""
metadata.py  —  file metadata extractors for everything_search.py
==================================================================
All functions swallow exceptions internally and return None fields
on failure rather than propagating errors.

Optional dependencies (each degrades gracefully):
    pip install mutagen Pillow opencv-python-headless openpyxl pymupdf python-docx rarfile py7zr
"""

from __future__ import annotations

import datetime
import email as _email_mod
import json
import os
import re
import sqlite3
import struct
import traceback
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

_DEBUG = False   # set True to print extraction errors to stderr

def _log(label: str, exc: Exception) -> None:
    if _DEBUG:
        import sys
        print(f"[metadata:{label}] {type(exc).__name__}: {exc}", file=sys.stderr)
        if _DEBUG > 1:
            traceback.print_exc(file=sys.stderr)

# ---------------------------------------------------------------------------
# Optional imports
# ---------------------------------------------------------------------------
try:
    import mutagen
    _MUTAGEN = True
except ImportError:
    _MUTAGEN = False

try:
    from PIL import Image as _PILImage
    _PIL = True
except ImportError:
    _PIL = False

try:
    import cv2 as _CV2
    _CV2_OK = True
except ImportError:
    _CV2_OK = False

try:
    import openpyxl as _OPENPYXL
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

try:
    import fitz as _FITZ
    _FITZ_OK = True
except ImportError:
    _FITZ_OK = False

try:
    import docx as _DOCX_LIB
    _DOCX_LIB_OK = True
except ImportError:
    _DOCX_LIB_OK = False

try:
    import csv as _CSV_MOD
    _CSV_OK = True
except ImportError:
    _CSV_OK = False

try:
    import rarfile as _RARFILE
    _RAR_OK = True
except ImportError:
    _RAR_OK = False

try:
    import py7zr as _PY7ZR
    _7Z_OK = True
except ImportError:
    _7Z_OK = False


# ===========================================================================
# HELPERS
# ===========================================================================

def _read_bytes(path: str, count: int, offset: int = 0) -> bytes:
    with open(path, 'rb') as f:
        f.seek(offset)
        return f.read(count)


def _fmt_duration(seconds: Optional[float]) -> Optional[str]:
    if seconds is None:
        return None
    s = int(seconds)
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    return f"{h}:{m:02d}:{sec:02d}" if h else f"{m}:{sec:02d}"


from utils import human_size as _human_size


# ===========================================================================
# IMAGE
# ===========================================================================

def image_info(path: str, ext: str) -> Dict[str, Any]:
    """Returns: dimensions (w,h), color_mode, bit_depth, dpi."""
    out: Dict[str, Any] = {'dimensions': None, 'color_mode': None,
                            'bit_depth': None, 'dpi': None}
    if _PIL:
        try:
            with _PILImage.open(path) as img:
                out['dimensions'] = img.size
                out['color_mode'] = img.mode
                out['bit_depth']  = _PIL_mode_depth(img.mode)
                dpi = img.info.get('dpi') or img.info.get('jfif_density')
                if dpi and isinstance(dpi, tuple) and dpi[0]:
                    out['dpi'] = f"{int(dpi[0])} DPI"
            return out
        except Exception as exc:
            _log('image_info/PIL', exc)
    out['dimensions'] = _image_dims_stdlib(path, ext)
    return out


def _PIL_mode_depth(mode: str) -> Optional[int]:
    return {'1': 1, 'L': 8, 'P': 8, 'RGB': 24, 'RGBA': 32,
            'CMYK': 32, 'YCbCr': 24, 'LAB': 24, 'HSV': 24,
            'I': 32, 'F': 32, 'LA': 16, 'PA': 16, 'RGBa': 32,
            'La': 16, 'I;16': 16, 'I;16B': 16}.get(mode)


def _image_dims_stdlib(path: str, ext: str) -> Optional[Tuple[int, int]]:
    e = ext.lower()
    try:
        with open(path, 'rb') as f:
            hdr = f.read(32)

        if e == 'png' and hdr[:8] == b'\x89PNG\r\n\x1a\n':
            w, h = struct.unpack('>II', hdr[16:24])
            return (w, h)

        if e in ('jpg', 'jpeg', 'jfif', 'jpe'):
            with open(path, 'rb') as f:
                data = f.read()
            i = 0
            while i < len(data) - 4:
                if data[i] != 0xFF:
                    break
                marker = data[i + 1]
                if marker in (0xC0, 0xC1, 0xC2):
                    h = struct.unpack('>H', data[i+5:i+7])[0]
                    w = struct.unpack('>H', data[i+7:i+9])[0]
                    return (w, h)
                if marker in (0xD8, 0xD9):
                    i += 2; continue
                seg_len = struct.unpack('>H', data[i+2:i+4])[0]
                i += 2 + seg_len
            return None

        if e == 'gif' and hdr[:6] in (b'GIF87a', b'GIF89a'):
            w, h = struct.unpack('<HH', hdr[6:10])
            return (w, h)

        if e == 'bmp' and hdr[:2] == b'BM':
            w, h = struct.unpack('<ii', hdr[18:26])
            return (abs(w), abs(h))

        if e == 'webp' and hdr[:4] == b'RIFF' and hdr[8:12] == b'WEBP':
            chunk = hdr[12:16]
            if chunk == b'VP8 ' and len(hdr) >= 30:
                w = struct.unpack('<H', hdr[26:28])[0] & 0x3FFF
                h = struct.unpack('<H', hdr[28:30])[0] & 0x3FFF
                return (w, h)
            if chunk == b'VP8L' and len(hdr) >= 25:
                bits = struct.unpack('<I', hdr[21:25])[0]
                w = (bits & 0x3FFF) + 1
                h = ((bits >> 14) & 0x3FFF) + 1
                return (w, h)

        if e in ('tif', 'tiff'):
            endian = '<' if hdr[:2] == b'II' else '>'
            offset = struct.unpack(endian + 'I', hdr[4:8])[0]
            with open(path, 'rb') as f:
                f.seek(offset)
                n = struct.unpack(endian + 'H', f.read(2))[0]
                w = h = None
                for _ in range(n):
                    entry = f.read(12)
                    if len(entry) < 12: break
                    tag = struct.unpack(endian + 'H', entry[:2])[0]
                    typ = struct.unpack(endian + 'H', entry[2:4])[0]
                    val_bytes = entry[8:12]
                    val = (struct.unpack(endian+'H', val_bytes[:2])[0]
                           if typ == 3 else
                           struct.unpack(endian+'I', val_bytes)[0])
                    if tag == 256: w = val
                    if tag == 257: h = val
            return (w, h) if w and h else None

        # HDR (Radiance RGBE)
        if e in ('hdr', 'rgbe'):
            with open(path, 'rb') as f:
                raw = f.read(512)
            # Header ends with blank line, then "-Y H +X W\n"
            try:
                text = raw.decode('latin-1', errors='replace')
                m = re.search(r'-Y\s+(\d+)\s+\+X\s+(\d+)', text)
                if m:
                    return (int(m.group(2)), int(m.group(1)))
            except Exception:
                pass

    except Exception as exc:
        _log('image_dims_stdlib', exc)
    return None


# ===========================================================================
# VIDEO
# ===========================================================================

def video_info(path: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {'dimensions': None, 'length_s': None,
                            'fps': None, 'codec': None}
    if not _CV2_OK:
        return out
    try:
        cap = _CV2.VideoCapture(path)
        if cap.isOpened():
            w   = int(cap.get(_CV2.CAP_PROP_FRAME_WIDTH))
            h   = int(cap.get(_CV2.CAP_PROP_FRAME_HEIGHT))
            fps = cap.get(_CV2.CAP_PROP_FPS)
            fc  = cap.get(_CV2.CAP_PROP_FRAME_COUNT)
            fourcc = int(cap.get(_CV2.CAP_PROP_FOURCC))
            codec = ''.join(chr((fourcc >> (8*i)) & 0xFF) for i in range(4)).strip('\x00 ')
            out['dimensions'] = (w, h) if w and h else None
            out['fps']        = round(fps, 3) if fps else None
            out['codec']      = codec or None
            if fps and fc:
                out['length_s'] = fc / fps
        cap.release()
    except Exception as exc:
        _log('video_info', exc)
    return out


# ===========================================================================
# AUDIO
# ===========================================================================

def audio_tags(path: str) -> Dict[str, Any]:
    keys = ('length_s','album','album_artist','artist','title','track',
            'year','genre','bitrate','sample_rate','channels')
    out: Dict[str, Any] = {k: None for k in keys}
    if not _MUTAGEN:
        return out
    try:
        audio = mutagen.File(path, easy=True)
        if audio is None:
            return out
        info = audio.info
        out['length_s']    = getattr(info, 'length', None)
        out['bitrate']     = getattr(info, 'bitrate', None)
        out['sample_rate'] = getattr(info, 'sample_rate', None)
        out['channels']    = getattr(info, 'channels', None)
        def _t(k):
            v = audio.get(k)
            return v[0] if v else None
        out['album']        = _t('album')
        out['album_artist'] = _t('albumartist')
        out['artist']       = _t('artist')
        out['title']        = _t('title')
        out['track']        = _t('tracknumber')
        out['year']         = _t('date') or _t('year')
        out['genre']        = _t('genre')
    except Exception as exc:
        _log('audio_tags', exc)
    return out


# ===========================================================================
# TEXT / CODE  —  line count
# ===========================================================================

# Extensions that get line counts
TEXT_EXTS = frozenset(
    'txt md rst log ini cfg conf toml yaml yml env '
    'py pyw pyi '
    'js mjs ts jsx tsx '
    'c h cc cpp cxx hh hpp hxx '
    'cs java rs go swift kt kts '
    'rb pl lua sh bash zsh fish '
    'jl '          # Julia
    'r R '         # R
    'f f90 f95 '   # Fortran
    'sql '
    'html htm xml xhtml svg '
    'css scss sass less '
    'json jsonc '
    'vb bas '
    'asm s '
    'cmake make makefile '
    'bat cmd ps1 psm1 '
    'php '
    'ex exs '      # Elixir
    'clj cljs '    # Clojure
    'hs '          # Haskell
    'erl hrl '     # Erlang
    'ml mli '      # OCaml
    'nim '
    'zig '
    'dart '
    'v vsh '       # V lang
    'proto '       # protobuf
    'tf tfvars '   # Terraform
    'gradle '
    'dockerfile '
    'gitignore '
    'editorconfig '
    'csv tsv '
    'properties '
    'lock '        # package-lock, Cargo.lock, etc.
    'rtf '
    'srt vtt sub ' # subtitle files
    .split()
)

def text_line_count(path: str, ext: str) -> Optional[str]:
    """Count lines in any text/code file."""
    try:
        # Try UTF-8 first, fall back to latin-1
        for encoding in ('utf-8', 'utf-16', 'latin-1'):
            try:
                with open(path, 'r', encoding=encoding, errors='strict') as f:
                    lines = sum(1 for _ in f)
                return f"{lines:,} lines"
            except (UnicodeDecodeError, UnicodeError):
                continue
    except Exception as exc:
        _log('text_line_count', exc)
    return None


# ===========================================================================
# ARCHIVE
# ===========================================================================

def archive_info(path: str, ext: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        'file_count': None, 'dir_count': None,
        'total_uncompressed': None, 'compression_method': None,
        'comment': None, 'encrypted': None, 'nested_archives': None,
    }
    e = ext.lower()
    try:
        if e in ('zip','jar','war','ear','apk','whl','cbz','epub','xlsx',
                 'docx','pptx','odt','ods','odp'):
            return _zip_info(path)
        if e in ('gz',) and not (path.endswith('.tar.gz') or path.endswith('.tgz')):
            return _gz_info(path)
        if e in ('tar','tgz') or path.endswith('.tar.gz') or path.endswith('.tar.bz2'):
            return _tar_info(path)
        if e in ('rar','cbr') and _RAR_OK:
            return _rar_info(path)
        if e == '7z' and _7Z_OK:
            return _7z_info(path)
    except Exception as exc:
        _log(f'archive_info/{e}', exc)
    return out


def _zip_info(path: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        'file_count': None, 'dir_count': None,
        'total_uncompressed': None, 'compression_method': None,
        'comment': None, 'encrypted': None, 'nested_archives': None,
    }
    _METHODS = {
        0:'Stored', 1:'Shrunk', 6:'Implode', 8:'Deflate',
        9:'Deflate64', 12:'BZip2', 14:'LZMA', 93:'Zstandard',
        95:'XZ', 99:'AES-encrypted',
    }
    _ARCH_EXTS = frozenset('zip rar 7z gz tar tgz bz2 jar'.split())
    try:
        with zipfile.ZipFile(path, 'r') as z:
            entries = z.infolist()
            files = [e for e in entries if not e.filename.endswith('/')]
            dirs  = [e for e in entries if e.filename.endswith('/')]
            methods = {_METHODS.get(e.compress_type, str(e.compress_type)) for e in files}
            encrypted = sum(1 for e in files if e.flag_bits & 0x1)
            nested   = sum(1 for e in files
                           if Path(e.filename).suffix.lower().lstrip('.') in _ARCH_EXTS)
            comment  = z.comment.decode('utf-8', errors='replace').strip() if z.comment else None
            out['file_count']         = len(files)
            out['dir_count']          = len(dirs)
            out['total_uncompressed'] = _human_size(sum(e.file_size for e in files))
            out['compression_method'] = ', '.join(sorted(methods)) or None
            out['comment']            = comment or None
            out['encrypted']          = (f"{encrypted} encrypted" if encrypted else 'No')
            out['nested_archives']    = nested or None
    except Exception as exc:
        _log('_zip_info', exc)
    return out


def _tar_info(path: str) -> Dict[str, Any]:
    import tarfile
    out: Dict[str, Any] = {
        'file_count': None, 'dir_count': None,
        'total_uncompressed': None, 'compression_method': None,
        'comment': None, 'encrypted': None, 'nested_archives': None,
    }
    try:
        if   path.endswith('.gz') or path.endswith('.tgz'): comp = 'gz'
        elif path.endswith('.bz2'):                         comp = 'bz2'
        elif path.endswith('.xz'):                          comp = 'xz'
        else:                                               comp = ''
        method = {'gz':'GZip','bz2':'BZip2','xz':'XZ','':'None'}.get(comp,'TAR')
        with tarfile.open(path, f'r:{comp}' if comp else 'r') as t:
            members = t.getmembers()
            files   = [m for m in members if m.isfile()]
            dirs    = [m for m in members if m.isdir()]
            out['file_count']         = len(files)
            out['dir_count']          = len(dirs)
            out['total_uncompressed'] = _human_size(sum(m.size for m in files))
            out['compression_method'] = method
            out['encrypted']          = 'No'
    except Exception as exc:
        _log('_tar_info', exc)
    return out


def _gz_info(path: str) -> Dict[str, Any]:
    import gzip
    out: Dict[str, Any] = {
        'file_count': 1, 'dir_count': 0,
        'total_uncompressed': None, 'compression_method': 'GZip',
        'comment': None, 'encrypted': 'No', 'nested_archives': None,
    }
    try:
        with gzip.open(path, 'rb') as f:
            out['total_uncompressed'] = _human_size(len(f.read()))
    except Exception as exc:
        _log('_gz_info', exc)
    return out


def _rar_info(path: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        'file_count': None, 'dir_count': None,
        'total_uncompressed': None, 'compression_method': 'RAR',
        'comment': None, 'encrypted': None, 'nested_archives': None,
    }
    try:
        with _RARFILE.RarFile(path) as rf:
            infos = rf.infolist()
            files = [i for i in infos if not i.is_dir()]
            dirs  = [i for i in infos if i.is_dir()]
            out['file_count']         = len(files)
            out['dir_count']          = len(dirs)
            out['total_uncompressed'] = _human_size(sum(i.file_size for i in files))
            out['encrypted']          = 'Yes' if rf.needs_password() else 'No'
            out['comment']            = rf.comment or None
    except Exception as exc:
        _log('_rar_info', exc)
    return out


def _7z_info(path: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        'file_count': None, 'dir_count': None,
        'total_uncompressed': None, 'compression_method': '7-Zip',
        'comment': None, 'encrypted': None, 'nested_archives': None,
    }
    try:
        with _PY7ZR.SevenZipFile(path, mode='r') as z:
            files   = z.list()
            regular = [f for f in files if not f.is_directory]
            dirs    = [f for f in files if f.is_directory]
            out['file_count']         = len(regular)
            out['dir_count']          = len(dirs)
            out['total_uncompressed'] = _human_size(
                sum(f.uncompressed for f in regular if f.uncompressed))
            out['encrypted']          = 'Yes' if z.needs_password() else 'No'
    except Exception as exc:
        _log('_7z_info', exc)
    return out


# ===========================================================================
# DOCUMENTS
# ===========================================================================

def doc_length(path: str, ext: str) -> Optional[str]:
    e = ext.lower()
    try:
        if e == 'pdf':
            n = _pdf_pages(path)
            return f"{n} pages" if n else None
        if e in ('docx','docm','dotx','dotm'):
            return _docx_length(path)
        if e == 'odt':
            return _odt_length(path)
        if e == 'epub':
            return _epub_length(path)
        if e == 'rtf':
            return _rtf_length(path)
        if e == 'txt':
            return text_line_count(path, ext)
    except Exception as exc:
        _log(f'doc_length/{e}', exc)
    return None


def _pdf_pages(path: str) -> Optional[int]:
    # Strategy 1: PyMuPDF (most reliable, handles compressed streams)
    if _FITZ_OK:
        try:
            return _FITZ.open(path).page_count
        except Exception as exc:
            _log('_pdf_pages/fitz', exc)

    # Strategy 2: find the root /Pages node with the highest /Count value
    # The document root /Pages dict contains /Count = total page count.
    # We take the maximum value found (handles linearized PDFs and nested Pages trees).
    try:
        with open(path, 'rb') as f:
            # Read in chunks to handle large PDFs without loading everything
            chunks = []
            while True:
                chunk = f.read(65536)
                if not chunk:
                    break
                chunks.append(chunk)
        data = b''.join(chunks)

        counts = re.findall(rb'/Count\s+(\d+)', data)
        if counts:
            return max(int(c) for c in counts)
    except Exception as exc:
        _log('_pdf_pages/count_regex', exc)

    # Strategy 3: count /Type /Page (misses compressed streams but catches simple PDFs)
    try:
        count = len(re.findall(rb'/Type\s*/Page[^s\w]', data))
        return count or None
    except Exception as exc:
        _log('_pdf_pages/type_page', exc)

    return None


def _docx_length(path: str) -> Optional[str]:
    try:
        with zipfile.ZipFile(path, 'r') as z:
            names = z.namelist()
            # Strategy 1: docProps/app.xml → Pages (written by Word on save)
            if 'docProps/app.xml' in names:
                try:
                    root = ET.parse(z.open('docProps/app.xml')).getroot()
                    for elem in root.iter():
                        if elem.tag.endswith('}Pages') or elem.tag == 'Pages':
                            p = int(elem.text or 0)
                            if p > 0:
                                return f"{p} pages"
                except Exception as exc:
                    _log('_docx_length/app.xml', exc)
            # Strategy 2: page break markers
            if 'word/document.xml' in names:
                try:
                    content = z.read('word/document.xml')
                    rendered = content.count(b'w:lastRenderedPageBreak')
                    explicit = content.count(b'w:type="page"')
                    sect     = content.count(b'<w:sectPr')
                    pages    = max(rendered, explicit + 1, sect)
                    if pages > 1:
                        return f"~{pages} pages (estimated)"
                    paras = content.count(b'<w:p ')
                    return f"{paras:,} paragraphs"
                except Exception as exc:
                    _log('_docx_length/document.xml', exc)
    except Exception as exc:
        _log('_docx_length', exc)
    return None


def _odt_length(path: str) -> Optional[str]:
    try:
        with zipfile.ZipFile(path, 'r') as z:
            if 'meta.xml' in z.namelist():
                root = ET.parse(z.open('meta.xml')).getroot()
                for elem in root.iter():
                    if elem.tag.endswith('}page-count'):
                        return f"{elem.text} pages"
    except Exception as exc:
        _log('_odt_length', exc)
    return None


def _epub_length(path: str) -> Optional[str]:
    try:
        with zipfile.ZipFile(path, 'r') as z:
            opf_path = None
            if 'META-INF/container.xml' in z.namelist():
                root = ET.parse(z.open('META-INF/container.xml')).getroot()
                for elem in root.iter():
                    if elem.tag.endswith('}rootfile'):
                        opf_path = elem.get('full-path')
                        break
            if opf_path and opf_path in z.namelist():
                root = ET.parse(z.open(opf_path)).getroot()
                chapters = sum(1 for e in root.iter() if e.tag.endswith('}itemref'))
                if chapters:
                    return f"{chapters} chapters"
    except Exception as exc:
        _log('_epub_length', exc)
    return None


def _rtf_length(path: str) -> Optional[str]:
    try:
        with open(path, 'rb') as f:
            data = f.read(1024 * 1024)
        pages = data.count(b'\\page') + 1
        return f"~{pages} pages (estimated)"
    except Exception as exc:
        _log('_rtf_length', exc)
    return None


# ===========================================================================
# SPREADSHEET
# ===========================================================================

def spreadsheet_info(path: str, ext: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {'dimensions': None, 'sheet_count': None}
    e = ext.lower()
    try:
        if e == 'csv' and _CSV_OK:
            with open(path, newline='', encoding='utf-8', errors='replace') as f:
                rows = list(_CSV_MOD.reader(f))
            out['dimensions']  = (len(rows), max((len(r) for r in rows), default=0))
            out['sheet_count'] = 1
            return out

        if e in ('xlsx','xlsm','xltx','xltm','xlsb'):
            if _OPENPYXL_OK:
                try:
                    wb = _OPENPYXL.load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    out['dimensions']  = (ws.max_row or 0, ws.max_column or 0)
                    out['sheet_count'] = len(wb.sheetnames)
                    return out
                except Exception as exc:
                    _log('spreadsheet_info/openpyxl', exc)
            # stdlib XML fallback
            try:
                with zipfile.ZipFile(path, 'r') as z:
                    if 'xl/workbook.xml' in z.namelist():
                        root = ET.parse(z.open('xl/workbook.xml')).getroot()
                        out['sheet_count'] = sum(
                            1 for e in root.iter() if e.tag.endswith('}sheet'))
                    sheet_files = sorted(
                        n for n in z.namelist()
                        if re.match(r'xl/worksheets/sheet\d+\.xml', n))
                    if sheet_files:
                        root = ET.parse(z.open(sheet_files[0])).getroot()
                        for elem in root.iter():
                            if elem.tag.endswith('}dimension'):
                                m = re.match(r'[A-Z]+\d+:([A-Z]+)(\d+)', elem.get('ref',''))
                                if m:
                                    col_str, row_str = m.groups()
                                    cols = sum((ord(c)-64)*(26**i)
                                               for i,c in enumerate(reversed(col_str)))
                                    out['dimensions'] = (int(row_str), cols)
            except Exception as exc:
                _log('spreadsheet_info/stdlib', exc)
    except Exception as exc:
        _log('spreadsheet_info', exc)
    return out


# ===========================================================================
# PRESENTATION
# ===========================================================================

def presentation_info(path: str, ext: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {'slide_count': None, 'notes_count': None}
    try:
        with zipfile.ZipFile(path, 'r') as z:
            names = z.namelist()
            out['slide_count'] = sum(
                1 for n in names if re.match(r'ppt/slides/slide\d+\.xml', n)) or None
            out['notes_count'] = sum(
                1 for n in names if re.match(r'ppt/notesSlides/notesSlide\d+\.xml', n)) or None
    except Exception as exc:
        _log('presentation_info', exc)
    return out


# ===========================================================================
# SHORTCUT (.lnk)  —  MS-SHLLINK binary format
# ===========================================================================

_LNK_HEADER_SZ = 0x4C
_LNK_GUID      = b'\x01\x14\x02\x00\x00\x00\x00\x00\xC0\x00\x00\x00\x00\x00\x00\x46'

_LNK_HAS_IDLIST   = 0x00000001
_LNK_HAS_LINKINFO = 0x00000002
_LNK_HAS_NAME     = 0x00000004
_LNK_HAS_RELPATH  = 0x00000008
_LNK_HAS_WORKDIR  = 0x00000010
_LNK_HAS_ARGS     = 0x00000020
_LNK_HAS_ICON     = 0x00000040
_LNK_IS_UNICODE   = 0x00000080
_LNK_HAS_ENVLOC   = 0x00000200


def lnk_info(path: str) -> Dict[str, Any]:
    """
    Parse a Windows .lnk (Shell Link) file.
    Returns: target, working_dir, arguments, icon, description,
             target_size, target_modified, hotkey, show_cmd.
    """
    out: Dict[str, Any] = {
        'target': None, 'working_dir': None, 'arguments': None,
        'icon': None, 'description': None, 'target_size': None,
        'target_modified': None, 'hotkey': None, 'show_cmd': None,
    }
    try:
        with open(path, 'rb') as f:
            data = f.read()

        if len(data) < _LNK_HEADER_SZ:
            return out
        if struct.unpack_from('<I', data, 0)[0] != _LNK_HEADER_SZ:
            return out
        if data[4:20] != _LNK_GUID:
            return out

        link_flags   = struct.unpack_from('<I', data, 20)[0]
        is_unicode   = bool(link_flags & _LNK_IS_UNICODE)

        target_write = struct.unpack_from('<Q', data, 28)[0]
        target_sz    = struct.unpack_from('<I', data, 52)[0]
        hotkey_raw   = struct.unpack_from('<H', data, 56)[0]
        show_cmd     = struct.unpack_from('<I', data, 60)[0]

        if target_sz:
            out['target_size'] = _human_size(target_sz)
        if target_write:
            out['target_modified'] = _lnk_filetime(target_write)
        if hotkey_raw:
            out['hotkey'] = _lnk_hotkey(hotkey_raw)
        out['show_cmd'] = {1:'Normal', 2:'Minimized', 3:'Maximized'}.get(show_cmd)

        pos = _LNK_HEADER_SZ

        # Skip IDList
        if link_flags & _LNK_HAS_IDLIST:
            idlist_sz = struct.unpack_from('<H', data, pos)[0]
            pos += 2 + idlist_sz

        # LinkInfo block  →  local target path
        if link_flags & _LNK_HAS_LINKINFO:
            li_size  = struct.unpack_from('<I', data, pos)[0]
            li_flags = struct.unpack_from('<I', data, pos + 4)[0]
            li_hdr_sz = struct.unpack_from('<I', data, pos + 8)[0]

            # Local path (ANSI then Unicode)
            if li_flags & 0x1:
                local_off = struct.unpack_from('<I', data, pos + 16)[0]
                if local_off:
                    out['target'] = _lnk_sz(data, pos + local_off)
                # Unicode local path (offset 28 in LinkInfo header if header >= 0x1C)
                if li_hdr_sz >= 0x1C:
                    uni_local_off = struct.unpack_from('<I', data, pos + 28)[0]
                    if uni_local_off:
                        uni = _lnk_wsz(data, pos + uni_local_off)
                        if uni:
                            out['target'] = uni   # prefer Unicode

            # UNC path for network targets
            if li_flags & 0x2 and not out['target']:
                net_off = struct.unpack_from('<I', data, pos + 20)[0]
                if net_off:
                    net_name_off = struct.unpack_from('<I', data, pos + net_off + 8)[0]
                    out['target'] = _lnk_sz(data, pos + net_off + net_name_off)
            pos += li_size

        # StringData section
        def _str() -> Optional[str]:
            nonlocal pos
            if pos + 2 > len(data):
                return None
            count = struct.unpack_from('<H', data, pos)[0]
            pos += 2
            if is_unicode:
                raw = data[pos:pos + count * 2]
                pos += count * 2
                return raw.decode('utf-16-le', errors='replace') or None
            else:
                raw = data[pos:pos + count]
                pos += count
                return raw.decode('latin-1', errors='replace') or None

        out['description'] = _str() if link_flags & _LNK_HAS_NAME    else None
        _str()                       if link_flags & _LNK_HAS_RELPATH  else None  # skip relpath
        out['working_dir'] = _str() if link_flags & _LNK_HAS_WORKDIR  else None
        out['arguments']   = _str() if link_flags & _LNK_HAS_ARGS     else None
        out['icon']        = _str() if link_flags & _LNK_HAS_ICON     else None

        # EnvironmentVariable ExtraData block (target with %env% vars)
        if not out['target'] and (link_flags & _LNK_HAS_ENVLOC):
            extra = data[pos:]
            for sig in (b'\x01\x00\x00\xA0', b'\x02\x00\x00\xA0'):
                idx = extra.find(sig)
                if idx != -1:
                    ansi = _lnk_sz(extra, idx + 4)
                    if ansi:
                        out['target'] = ansi
                    break

    except Exception as exc:
        _log('lnk_info', exc)
    return out


def _lnk_sz(data: bytes, offset: int) -> Optional[str]:
    try:
        end = data.index(b'\x00', offset)
        s = data[offset:end].decode('latin-1', errors='replace')
        return s or None
    except (ValueError, IndexError):
        return None


def _lnk_wsz(data: bytes, offset: int) -> Optional[str]:
    try:
        chars = []
        i = offset
        while i + 1 < len(data):
            wc = data[i:i+2]
            if wc == b'\x00\x00':
                break
            chars.append(wc)
            i += 2
        s = b''.join(chars).decode('utf-16-le', errors='replace')
        return s or None
    except Exception:
        return None


def _lnk_filetime(ticks: int) -> Optional[str]:
    _EPOCH_DIFF = (
        datetime.datetime(1970,1,1) - datetime.datetime(1601,1,1)
    ).total_seconds() * 10_000_000
    if ticks in (0, 2**64-1):
        return None
    try:
        dt = datetime.datetime.fromtimestamp((ticks - _EPOCH_DIFF) / 10_000_000)
        return dt.strftime('%d/%m/%Y %I:%M %p')
    except (OSError, ValueError):
        return None


def _lnk_hotkey(raw: int) -> Optional[str]:
    vk = raw & 0xFF
    if not vk:
        return None
    vk_map = {0x70:'F1',0x71:'F2',0x72:'F3',0x73:'F4',0x74:'F5',
              0x75:'F6',0x76:'F7',0x77:'F8',0x78:'F9',0x79:'F10',
              0x7A:'F11',0x7B:'F12'}
    key = vk_map.get(vk, chr(vk) if 0x41<=vk<=0x5A else f'VK{vk:02X}')
    mod = (raw >> 8) & 0xFF
    mods = (['Shift'] if mod&1 else []) + (['Ctrl'] if mod&2 else []) + (['Alt'] if mod&4 else [])
    return '+'.join(mods + [key]) if mods else key


# ===========================================================================
# CRYPTO / PGP / GPG / KEY / PEM
# ===========================================================================

_PGP_PACKET_TAGS: Dict[int, str] = {
    1:'Public-Key Encrypted Session Key', 2:'Signature',
    3:'Symmetric-Key Encrypted Session Key', 4:'One-Pass Signature',
    5:'Secret-Key', 6:'Public-Key', 7:'Secret-Subkey',
    8:'Compressed Data', 9:'Symmetrically Encrypted Data',
    10:'Marker', 11:'Literal Data', 12:'Trust', 13:'User ID',
    14:'Public-Subkey', 17:'User Attribute',
    18:'Sym. Encrypted Integrity Protected Data',
    19:'Modification Detection Code',
}

_PGP_KEY_ALGOS: Dict[int, str] = {
    1:'RSA (enc+sign)', 2:'RSA (enc)', 3:'RSA (sign)',
    16:'ElGamal', 17:'DSA', 18:'ECDH', 19:'ECDSA', 22:'EdDSA',
}

_PGP_ARMOR_MARKERS = {
    '-----BEGIN PGP PUBLIC KEY BLOCK-----':    'PGP Public Key',
    '-----BEGIN PGP PRIVATE KEY BLOCK-----':   'PGP Private Key',
    '-----BEGIN PGP MESSAGE-----':             'PGP Encrypted Message',
    '-----BEGIN PGP SIGNED MESSAGE-----':      'PGP Signed Message',
    '-----BEGIN PGP SIGNATURE-----':           'PGP Signature',
    '-----BEGIN CERTIFICATE-----':             'X.509 Certificate (PEM)',
    '-----BEGIN RSA PRIVATE KEY-----':         'RSA Private Key (PEM)',
    '-----BEGIN EC PRIVATE KEY-----':          'EC Private Key (PEM)',
    '-----BEGIN OPENSSH PRIVATE KEY-----':     'OpenSSH Private Key',
    '-----BEGIN SSH2 PUBLIC KEY-----':         'SSH2 Public Key',
}

_SSH_KEY_TYPES = {
    b'ssh-rsa':'RSA', b'ssh-dss':'DSA',
    b'ecdsa-sha2-nistp256':'ECDSA-256',
    b'ecdsa-sha2-nistp384':'ECDSA-384',
    b'ecdsa-sha2-nistp521':'ECDSA-521',
    b'ssh-ed25519':'Ed25519',
    b'sk-ssh-ed25519@openssh.com':'Ed25519-SK',
}


def crypto_info(path: str, ext: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        'key_type':None, 'algorithm':None, 'key_id':None,
        'fingerprint':None, 'created':None, 'expires':None,
        'user_id':None, 'packet_types':None, 'armored':None, 'comment':None,
    }
    try:
        with open(path, 'rb') as f:
            raw = f.read(65536)

        text = raw[:256].lstrip()
        if text.startswith(b'-----'):
            out['armored'] = True
            first_line = text.split(b'\n')[0].decode('ascii', errors='replace').strip()
            for marker, label in _PGP_ARMOR_MARKERS.items():
                if first_line == marker:
                    out['key_type'] = label
                    break
            for line in raw[:1024].decode('ascii', errors='replace').splitlines():
                line = line.strip()
                if line.startswith('Comment:'):
                    out['comment'] = line[8:].strip()
                if not line:
                    break
            if out['key_type'] is None:
                for ktype, label in _SSH_KEY_TYPES.items():
                    if ktype in raw:
                        out['key_type'] = f'SSH Public Key ({label})'
                        parts = raw.decode('ascii', errors='replace').split()
                        if len(parts) >= 3:
                            out['comment'] = parts[2]
                        break
            if out['key_type'] and 'PGP' in (out['key_type'] or ''):
                import base64
                body_lines = []
                in_body = False
                for line in raw.decode('ascii', errors='replace').splitlines():
                    if in_body:
                        if line.startswith('-----END') or line.startswith('='):
                            break
                        body_lines.append(line.strip())
                    elif line == '':
                        in_body = True
                if body_lines:
                    try:
                        decoded = base64.b64decode(''.join(body_lines))
                        out.update(_parse_pgp_packets(decoded))
                    except Exception as exc:
                        _log('crypto_info/base64_decode', exc)
        else:
            out['armored'] = False
            if raw and (raw[0] & 0x80):
                out.update(_parse_pgp_packets(raw))
    except Exception as exc:
        _log('crypto_info', exc)
    return out


def _parse_pgp_packets(data: bytes) -> Dict[str, Any]:
    result: Dict[str, Any] = {
        'key_type':None, 'algorithm':None, 'key_id':None,
        'created':None, 'expires':None, 'user_id':None, 'packet_types':None,
    }
    packet_tags: List[str] = []
    pos = 0
    try:
        while pos < len(data) - 2:
            byte = data[pos]
            if not (byte & 0x80):
                break
            new_fmt = bool(byte & 0x40)
            if new_fmt:
                tag = byte & 0x3F
                pos += 1
                if pos >= len(data): break
                lb = data[pos]; pos += 1
                if lb < 192:
                    body_len = lb
                elif lb < 224:
                    body_len = ((lb - 192) << 8) + data[pos] + 192; pos += 1
                elif lb == 255:
                    body_len = struct.unpack_from('>I', data, pos)[0]; pos += 4
                else:
                    body_len = 1 << (lb & 0x1F)
            else:
                tag = (byte & 0x3C) >> 2
                lt  = byte & 0x03; pos += 1
                if lt == 0: body_len = data[pos]; pos += 1
                elif lt == 1: body_len = struct.unpack_from('>H',data,pos)[0]; pos += 2
                elif lt == 2: body_len = struct.unpack_from('>I',data,pos)[0]; pos += 4
                else: body_len = len(data) - pos

            tag_name = _PGP_PACKET_TAGS.get(tag, f'Unknown({tag})')
            if tag_name not in packet_tags:
                packet_tags.append(tag_name)

            body = data[pos:pos+body_len]
            if tag in (6,14) and len(body) >= 6 and result['algorithm'] is None:
                if body[0] == 4:
                    ts   = struct.unpack('>I', body[1:5])[0]
                    algo = body[5]
                    result['algorithm'] = _PGP_KEY_ALGOS.get(algo, f'Algo {algo}')
                    result['key_type']  = 'PGP Public Key' if tag==6 else 'PGP Secret Key'
                    try:
                        result['created'] = datetime.datetime.fromtimestamp(ts).strftime('%d/%m/%Y')
                    except Exception:
                        pass
            if tag == 2 and len(body) >= 6 and result['expires'] is None:
                _parse_sig_subpackets(body, result)
            if tag == 13:
                uid = body.decode('utf-8', errors='replace').strip()
                if uid:
                    result['user_id'] = uid
            pos += body_len
    except Exception as exc:
        _log('_parse_pgp_packets', exc)
    if packet_tags:
        result['packet_types'] = ', '.join(packet_tags)
    return result


def _parse_sig_subpackets(body: bytes, result: Dict[str, Any]) -> None:
    try:
        if len(body) < 6 or body[0] != 4:
            return
        hashed_count = struct.unpack('>H', body[4:6])[0]
        pos = 6; end = pos + hashed_count
        while pos < end and pos < len(body):
            spkt_len  = body[pos]; pos += 1
            if spkt_len == 0: break
            spkt_type = body[pos]
            spkt_body = body[pos+1:pos+spkt_len]
            if spkt_type == 9 and len(spkt_body) >= 4:
                exp = struct.unpack('>I', spkt_body[:4])[0]
                if exp and result.get('created'):
                    try:
                        cd = datetime.datetime.strptime(result['created'], '%d/%m/%Y')
                        result['expires'] = (cd + datetime.timedelta(seconds=exp)).strftime('%d/%m/%Y')
                    except Exception:
                        result['expires'] = f"{exp//(86400*365):.0f} years after creation"
            if spkt_type == 16 and len(spkt_body) >= 8 and not result.get('key_id'):
                result['key_id'] = spkt_body[:8].hex().upper()
            pos += spkt_len
    except Exception as exc:
        _log('_parse_sig_subpackets', exc)


# ===========================================================================
# SQLITE / DATABASE
# ===========================================================================

def sqlite_info(path: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        'table_count':None, 'tables':None, 'page_size':None,
        'page_count':None, 'encoding':None, 'application_id':None, 'user_version':None,
    }
    try:
        magic = _read_bytes(path, 16)
        if magic != b'SQLite format 3\x00':
            return out
    except Exception:
        return out
    try:
        hdr = _read_bytes(path, 100)
        out['page_size']      = struct.unpack('>H', hdr[16:18])[0]
        out['page_count']     = struct.unpack('>I', hdr[28:32])[0]
        enc                   = struct.unpack('>I', hdr[56:60])[0]
        out['encoding']       = {1:'UTF-8',2:'UTF-16-LE',3:'UTF-16-BE'}.get(enc, f'Enc{enc}')
        out['user_version']   = struct.unpack('>I', hdr[60:64])[0]
        out['application_id'] = struct.unpack('>I', hdr[68:72])[0] or None
        conn = sqlite3.connect(f'file:{path}?mode=ro', uri=True, timeout=2.0)
        tables = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
        ).fetchall()
        out['table_count'] = len(tables)
        info = []
        for (tname,) in tables:
            try:
                count = conn.execute(f'SELECT COUNT(*) FROM "{tname}"').fetchone()[0]
                info.append({'name': tname, 'rows': count})
            except sqlite3.Error:
                info.append({'name': tname, 'rows': None})
        out['tables'] = info
        conn.close()
    except Exception as exc:
        _log('sqlite_info', exc)
    return out


# ===========================================================================
# TORRENT
# ===========================================================================

def torrent_info(path: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        'name':None, 'file_count':None, 'total_size':None,
        'tracker':None, 'trackers':None, 'comment':None,
        'created_by':None, 'creation_date':None, 'private':None,
    }
    try:
        with open(path, 'rb') as f:
            raw = f.read()

        # Guard: must start with 'd' (bencode dict)
        if not raw or chr(raw[0]) != 'd':
            return out

        torrent, _ = _bdecode(raw)
        if not isinstance(torrent, dict):
            return out

        def _s(v: Any) -> Optional[str]:
            if isinstance(v, bytes):
                return v.decode('utf-8', errors='replace')
            return str(v) if v is not None else None

        announce = torrent.get('announce')
        if announce:
            out['tracker'] = _s(announce)

        all_trackers: set = set()
        if out['tracker']:
            all_trackers.add(out['tracker'])
        for tier in torrent.get('announce-list', []):
            if isinstance(tier, list):
                for t in tier:
                    s = _s(t)
                    if s: all_trackers.add(s)
        if len(all_trackers) > 1:
            out['trackers'] = f"{len(all_trackers)} trackers"

        out['comment']    = _s(torrent.get('comment')) or None
        out['created_by'] = _s(torrent.get('created by')) or None

        creation = torrent.get('creation date')
        if creation:
            try:
                dt = datetime.datetime.fromtimestamp(int(creation))
                out['creation_date'] = dt.strftime('%d/%m/%Y')
            except (OSError, ValueError):
                pass

        info = torrent.get('info', {})
        if not isinstance(info, dict):
            return out

        name = info.get('name')
        out['name']    = _s(name) or None
        out['private'] = 'Yes' if info.get('private') == 1 else 'No'

        files = info.get('files')
        if files and isinstance(files, list):
            out['file_count'] = len(files)
            out['total_size'] = _human_size(
                sum(f.get('length', 0) for f in files if isinstance(f, dict)))
        else:
            out['file_count'] = 1
            length = info.get('length')
            out['total_size'] = _human_size(length) if length else None

    except Exception as exc:
        _log('torrent_info', exc)
    return out


def _bdecode(data: bytes, idx: int = 0):
    if idx >= len(data):
        raise ValueError(f"Unexpected end at position {idx}")
    c = data[idx]
    ch = chr(c)
    if ch == 'd':
        idx += 1; result: Dict[str, Any] = {}
        while idx < len(data) and chr(data[idx]) != 'e':
            key, idx = _bdecode(data, idx)
            val, idx = _bdecode(data, idx)
            k = key.decode('utf-8', errors='replace') if isinstance(key, bytes) else str(key)
            result[k] = val
        return result, idx + 1
    elif ch == 'l':
        idx += 1; result_list: List[Any] = []
        while idx < len(data) and chr(data[idx]) != 'e':
            val, idx = _bdecode(data, idx)
            result_list.append(val)
        return result_list, idx + 1
    elif ch == 'i':
        end = data.index(b'e', idx + 1)
        return int(data[idx+1:end]), end + 1
    elif 48 <= c <= 57:
        colon = data.index(b':', idx)
        length = int(data[idx:colon])
        start  = colon + 1
        return data[start:start+length], start + length
    else:
        raise ValueError(f"Invalid bencode byte {c!r} at position {idx}")


# ===========================================================================
# FONT
# ===========================================================================

def font_info(path: str, ext: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        'family':None, 'subfamily':None, 'full_name':None,
        'version':None, 'postscript_name':None,
        'glyph_count':None, 'units_per_em':None,
    }
    e = ext.lower()
    try:
        with open(path, 'rb') as f:
            raw = f.read()

        sfnt_offset = 0

        if e == 'woff':
            if raw[:4] != b'wOFF':
                return out
            num_tables = struct.unpack('>H', raw[12:14])[0]
            out.update(_woff_name_table(raw, num_tables))
            return out

        if e == 'woff2':
            if raw[:4] == b'wOF2':
                num_tables = struct.unpack('>H', raw[12:14])[0]
                out['family'] = f"WOFF2 font ({num_tables} tables)"
            return out

        sfnt_tag = raw[:4]
        if sfnt_tag not in (b'\x00\x01\x00\x00', b'OTTO', b'true', b'typ1', b'ttcf'):
            return out

        if sfnt_tag == b'ttcf':
            num_fonts = struct.unpack('>I', raw[8:12])[0]
            sfnt_offset = struct.unpack('>I', raw[12:16])[0]
            out['family'] = f"TTC ({num_fonts} fonts)"

        num_tables = struct.unpack('>H', raw[sfnt_offset+4:sfnt_offset+6])[0]
        tdir_off   = sfnt_offset + 12
        name_off   = head_off = maxp_off = None
        name_len   = None

        for i in range(num_tables):
            entry = raw[tdir_off + i*16 : tdir_off + i*16 + 16]
            if len(entry) < 16: break
            tag    = entry[0:4]
            offset = struct.unpack('>I', entry[8:12])[0]
            length = struct.unpack('>I', entry[12:16])[0]
            if tag == b'name': name_off = offset; name_len = length
            elif tag == b'head': head_off = offset
            elif tag == b'maxp': maxp_off = offset

        if head_off is not None and head_off + 20 <= len(raw):
            out['units_per_em'] = struct.unpack('>H', raw[head_off+18:head_off+20])[0]
        if maxp_off is not None and maxp_off + 6 <= len(raw):
            out['glyph_count'] = struct.unpack('>H', raw[maxp_off+4:maxp_off+6])[0]
        if name_off is not None:
            out.update(_parse_name_table(raw, name_off))
    except Exception as exc:
        _log('font_info', exc)
    return out


def _parse_name_table(raw: bytes, offset: int) -> Dict[str, Any]:
    result: Dict[str, Any] = {}
    _NAME_IDS = {1:'family', 2:'subfamily', 4:'full_name', 5:'version', 6:'postscript_name'}
    try:
        count   = struct.unpack('>H', raw[offset+2:offset+4])[0]
        str_off = struct.unpack('>H', raw[offset+4:offset+6])[0]
        strs_start = offset + str_off
        for i in range(count):
            rec = raw[offset + 6 + i*12 : offset + 6 + i*12 + 12]
            if len(rec) < 12: break
            platform_id = struct.unpack('>H', rec[0:2])[0]
            name_id     = struct.unpack('>H', rec[6:8])[0]
            str_length  = struct.unpack('>H', rec[8:10])[0]
            str_offset  = struct.unpack('>H', rec[10:12])[0]
            if name_id not in _NAME_IDS or result.get(_NAME_IDS[name_id]):
                continue
            s_bytes = raw[strs_start+str_offset : strs_start+str_offset+str_length]
            s = (s_bytes.decode('utf-16-be', errors='replace') if platform_id == 3
                 else s_bytes.decode('latin-1', errors='replace')).strip('\x00')
            if s:
                result[_NAME_IDS[name_id]] = s
    except Exception as exc:
        _log('_parse_name_table', exc)
    return result


def _woff_name_table(raw: bytes, num_tables: int) -> Dict[str, Any]:
    import zlib
    try:
        for i in range(num_tables):
            entry = raw[44 + i*20 : 44 + i*20 + 20]
            if len(entry) < 20: break
            if entry[0:4] != b'name': continue
            offset   = struct.unpack('>I', entry[4:8])[0]
            comp_len = struct.unpack('>I', entry[8:12])[0]
            orig_len = struct.unpack('>I', entry[12:16])[0]
            comp_data = raw[offset:offset+comp_len]
            table_data = zlib.decompress(comp_data) if comp_len < orig_len else comp_data
            return _parse_name_table(table_data, 0)
    except Exception as exc:
        _log('_woff_name_table', exc)
    return {}


# ===========================================================================
# EMAIL
# ===========================================================================

def email_info(path: str, ext: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        'subject':None, 'sender':None, 'recipients':None,
        'date':None, 'message_id':None,
        'attachment_count':None, 'content_type':None,
    }
    if ext.lower() != 'eml':
        return out
    try:
        with open(path, 'rb') as f:
            msg = _email_mod.message_from_bytes(f.read())
        out['subject']      = msg.get('Subject','').strip() or None
        out['sender']       = msg.get('From','').strip() or None
        out['recipients']   = msg.get('To','').strip() or None
        out['date']         = msg.get('Date','').strip() or None
        out['message_id']   = msg.get('Message-ID','').strip() or None
        out['content_type'] = msg.get('Content-Type','').split(';')[0].strip() or None
        attachments = sum(
            1 for part in msg.walk()
            if part.get_content_disposition() == 'attachment')
        out['attachment_count'] = attachments or None
    except Exception as exc:
        _log('email_info', exc)
    return out


# ===========================================================================
# 3D FILES
# ===========================================================================

def model_3d_info(path: str, ext: str) -> Optional[str]:
    e = ext.lower()
    parsers = {
        'obj':'_3d_obj', 'stl':'_3d_stl', 'ply':'_3d_ply',
        'gltf':'_3d_gltf', 'glb':'_3d_glb', 'dae':'_3d_dae',
        'fbx':'_3d_fbx', '3ds':'_3d_3ds', 'blend':'_3d_blend',
        'step':'_3d_step', 'stp':'_3d_step',
        'off':'_3d_off', 'amf':'_3d_amf', 'x3d':'_3d_x3d',
    }
    fn_name = parsers.get(e)
    if fn_name:
        fn = globals().get(fn_name)
        if fn:
            try:
                return fn(path)
            except Exception as exc:
                _log(f'model_3d_info/{e}', exc)
    return f"{e.upper()} file"


def _3d_obj(path: str) -> Optional[str]:
    counts: Dict[str,int] = {'v':0,'f':0,'vt':0,'vn':0,'g':0,'usemtl':0}
    try:
        with open(path,'r',errors='replace') as f:
            for line in f:
                tok = line.split()
                if tok and tok[0] in counts:
                    counts[tok[0]] += 1
        parts = [f"{counts['v']:,} verts", f"{counts['f']:,} faces"]
        if counts['vn']:     parts.append(f"{counts['vn']:,} normals")
        if counts['vt']:     parts.append(f"{counts['vt']:,} UVs")
        if counts['g']:      parts.append(f"{counts['g']} groups")
        if counts['usemtl']: parts.append(f"{counts['usemtl']} materials")
        return ', '.join(parts)
    except Exception as exc:
        _log('_3d_obj', exc); return None


def _3d_stl(path: str) -> Optional[str]:
    try:
        with open(path,'rb') as f:
            hdr = f.read(5)
        if hdr[:5].lower() == b'solid':
            with open(path,'r',errors='replace') as f:
                tris = sum(1 for l in f if l.strip().startswith('facet normal'))
            return f"{tris:,} triangles (ASCII STL)"
        else:
            with open(path,'rb') as f:
                f.seek(80)
                n = struct.unpack('<I',f.read(4))[0]
            return f"{n:,} triangles (binary STL)"
    except Exception as exc:
        _log('_3d_stl', exc); return None


def _3d_ply(path: str) -> Optional[str]:
    verts = faces = None
    try:
        with open(path,'rb') as f:
            for _ in range(200):
                line = f.readline().decode('ascii',errors='replace').strip()
                if line == 'end_header': break
                m = re.match(r'element vertex (\d+)', line)
                if m: verts = int(m.group(1))
                m = re.match(r'element face (\d+)', line)
                if m: faces = int(m.group(1))
        if verts is not None:
            parts = [f"{verts:,} verts"]
            if faces is not None: parts.append(f"{faces:,} faces")
            return ', '.join(parts)
    except Exception as exc:
        _log('_3d_ply', exc)
    return None


def _3d_gltf(path: str) -> Optional[str]:
    try:
        with open(path,'r',errors='replace') as f:
            return _gltf_summary(json.load(f), 'glTF')
    except Exception as exc:
        _log('_3d_gltf', exc); return None


def _3d_glb(path: str) -> Optional[str]:
    try:
        with open(path,'rb') as f:
            if f.read(4) != b'glTF': return None
            f.read(8)
            cl = struct.unpack('<I',f.read(4))[0]
            ct = f.read(4)
            if ct != b'JSON': return None
            return _gltf_summary(json.loads(f.read(cl).decode('utf-8',errors='replace')), 'glTF (binary)')
    except Exception as exc:
        _log('_3d_glb', exc); return None


def _gltf_summary(data: dict, label: str) -> str:
    version    = data.get('asset',{}).get('version','?')
    meshes     = len(data.get('meshes',[]))
    anims      = len(data.get('animations',[]))
    nodes      = len(data.get('nodes',[]))
    skins      = len(data.get('skins',[]))
    prims      = sum(len(m.get('primitives',[])) for m in data.get('meshes',[]))
    # Vertex count from accessors referenced by POSITION attributes
    accessors  = data.get('accessors', [])
    total_verts = 0
    total_faces = 0
    for mesh in data.get('meshes', []):
        for prim in mesh.get('primitives', []):
            attrs = prim.get('attributes', {})
            pos_idx = attrs.get('POSITION')
            if pos_idx is not None and pos_idx < len(accessors):
                total_verts += accessors[pos_idx].get('count', 0)
            idx_idx = prim.get('indices')
            if idx_idx is not None and idx_idx < len(accessors):
                total_faces += accessors[idx_idx].get('count', 0) // 3
    parts = [f"{label} {version}", f"{meshes} meshes"]
    if total_verts: parts.append(f"{total_verts:,} verts")
    if total_faces: parts.append(f"{total_faces:,} faces")
    if nodes:  parts.append(f"{nodes} nodes")
    if anims:  parts.append(f"{anims} animations")
    if skins:  parts.append(f"{skins} skins")
    return ', '.join(parts)


def _3d_dae(path: str) -> Optional[str]:
    try:
        root = ET.parse(path).getroot()
        ns   = re.match(r'\{(.+)\}', root.tag)
        p    = f'{{{ns.group(1)}}}' if ns else ''
        def _n(tag): return len(root.findall(f'.//{p}{tag}'))
        geoms=_n('geometry'); nodes=_n('node'); anims=_n('animation')
        skins=_n('skin');     images=_n('image')
        # vertex counts from float_array inside mesh/vertices
        verts = 0
        for fa in root.findall(f'.//{p}float_array'):
            parent = fa.get('id','')
            if 'position' in parent.lower() or 'vertex' in parent.lower():
                try:
                    verts += int(fa.get('count',0)) // 3
                except Exception:
                    pass
        parts = [f"COLLADA", f"{geoms} geometries", f"{nodes} nodes"]
        if verts:  parts.append(f"{verts:,} verts")
        if anims:  parts.append(f"{anims} animations")
        if skins:  parts.append(f"{skins} skins")
        if images: parts.append(f"{images} images")
        return ', '.join(parts)
    except Exception as exc:
        _log('_3d_dae', exc); return None


def _3d_fbx(path: str) -> Optional[str]:
    FBX_MAGIC = b'Kaydara FBX Binary  \x00\x1a\x00'
    try:
        with open(path,'rb') as f:
            hdr = f.read(27)
    except Exception as exc:
        _log('_3d_fbx/read', exc); return None
    if hdr[:23] == FBX_MAGIC[:23]:
        return _fbx_binary(path)
    return _fbx_ascii(path)


def _fbx_binary(path: str) -> Optional[str]:
    try:
        with open(path,'rb') as f:
            data = f.read()
        version  = struct.unpack('<I', data[23:27])[0]
        is_64bit = version >= 7500
        rec_fmt  = '<QQQ' if is_64bit else '<III'
        rec_sz   = 24 if is_64bit else 12
        fbx_ver  = f"{version//1000}.{(version%1000)//100}"

        node_counts: Dict[bytes,int] = {
            b'Geometry':0, b'Model':0, b'AnimStack':0,
            b'Deformer':0, b'Material':0, b'Texture':0,
        }
        total_verts = 0
        total_faces = 0

        # Scan for named sub-nodes inside the Objects block
        for obj_m in re.finditer(rb'Objects\x00', data[27:]):
            start = 27 + obj_m.end()
            pos = start
            for _ in range(200_000):
                if pos + rec_sz + 1 > len(data): break
                rec      = struct.unpack_from(rec_fmt, data, pos)
                end_off, n_props, prop_list_len = rec
                name_len = data[pos + rec_sz]
                name     = data[pos + rec_sz + 1 : pos + rec_sz + 1 + name_len]
                for key in node_counts:
                    if name == key:
                        node_counts[key] += 1
                if end_off == 0 or end_off <= pos: break
                pos = end_off
            break

        # Extract vertex and face counts from Vertices / PolygonVertexIndex arrays
        # These appear as typed array properties immediately following the node name
        # Array property format: type(1) + count(4) + encoding(4) + compressed_len(4)
        for arr_name, divisor in ((b'Vertices\x00', 3), (b'PolygonVertexIndex\x00', 1)):
            offset = 0
            while True:
                idx = data.find(arr_name, offset)
                if idx == -1: break
                prop_start = idx + len(arr_name)
                if prop_start + 9 <= len(data):
                    prop_type = chr(data[prop_start])
                    if prop_type in ('d', 'f', 'D', 'F', 'i', 'I', 'l', 'L'):
                        count = struct.unpack_from('<I', data, prop_start + 1)[0]
                        if arr_name == b'Vertices\x00':
                            total_verts += count // 3
                        else:
                            total_faces += count // 3   # approx (assumes triangles)
                offset = idx + 1
                if offset > len(data): break

        parts = [f"FBX {fbx_ver}"]
        if node_counts[b'Model']:    parts.append(f"{node_counts[b'Model']} models")
        if node_counts[b'Geometry']: parts.append(f"{node_counts[b'Geometry']} geometries")
        if total_verts:              parts.append(f"{total_verts:,} verts")
        if total_faces:              parts.append(f"{total_faces:,} faces")
        if node_counts[b'Material']: parts.append(f"{node_counts[b'Material']} materials")
        if node_counts[b'Texture']:  parts.append(f"{node_counts[b'Texture']} textures")
        if node_counts[b'AnimStack']:parts.append(f"{node_counts[b'AnimStack']} anim stacks")
        return ', '.join(parts)
    except Exception as exc:
        _log('_fbx_binary', exc); return "FBX (binary)"


def _fbx_ascii(path: str) -> Optional[str]:
    counts: Dict[str,int] = {
        'Geometry':0,'Model':0,'AnimStack':0,'Deformer':0,'Material':0,'Texture':0}
    total_verts = 0
    try:
        with open(path,'r',errors='replace') as f:
            for line in f:
                s = line.strip()
                for key in counts:
                    if s.startswith(f'{key}:') and '{' in s:
                        counts[key] += 1
                # Look for vertex array: Vertices: *N {
                m = re.match(r'Vertices:\s*\*(\d+)', s)
                if m:
                    total_verts += int(m.group(1)) // 3
        parts = [f"FBX (ASCII)"]
        if counts['Model']:    parts.append(f"{counts['Model']} models")
        if counts['Geometry']: parts.append(f"{counts['Geometry']} geometries")
        if total_verts:        parts.append(f"{total_verts:,} verts")
        if counts['Material']: parts.append(f"{counts['Material']} materials")
        return ', '.join(parts)
    except Exception as exc:
        _log('_fbx_ascii', exc); return "FBX (ASCII)"


def _3d_blend(path: str) -> Optional[str]:
    """
    Blender .blend — parse header + DNA1 block for scene info.
    Header: 'BLENDER' + ptr_size('-'=8,'_'=4) + endian('v'=little,'V'=big) + version(3)
    We scan for SDNA/DNA1 block and count SDObjects/Meshes if possible.
    """
    try:
        with open(path,'rb') as f:
            hdr = f.read(12)
        if hdr[:7] != b'BLENDER':
            return None
        ptr_ch   = chr(hdr[7])
        end_ch   = chr(hdr[8])
        version  = hdr[9:12].decode('ascii', errors='replace')
        ptr_size = 8 if ptr_ch == '-' else 4
        endian   = 'little' if end_ch == 'v' else 'big'
        end_sym  = '<' if endian == 'little' else '>'
        v        = f"{version[0]}.{version[1:]}"

        # Scan file blocks to count meshes and objects
        # Block header: code(4) + size(4) + old_ptr(ptr_size) + sdna_index(4) + count(4)
        blk_hdr_sz = 4 + 4 + ptr_size + 4 + 4
        mesh_count = obj_count = mat_count = 0

        with open(path,'rb') as f:
            f.seek(12)   # past header
            for _ in range(100_000):
                blk_hdr = f.read(blk_hdr_sz)
                if len(blk_hdr) < blk_hdr_sz: break
                code = blk_hdr[:4].rstrip(b'\x00').decode('ascii', errors='replace')
                blk_size = struct.unpack_from(end_sym+'I', blk_hdr, 4)[0]
                if code == 'ENDB': break
                if code == 'ME': mesh_count += 1
                if code == 'OB': obj_count  += 1
                if code == 'MA': mat_count  += 1
                f.seek(blk_size, 1)   # skip block data

        parts = [f"Blender {v}", f"{ptr_size*8}-bit", f"{endian}-endian"]
        if obj_count:  parts.append(f"{obj_count} objects")
        if mesh_count: parts.append(f"{mesh_count} meshes")
        if mat_count:  parts.append(f"{mat_count} materials")
        return ', '.join(parts)
    except Exception as exc:
        _log('_3d_blend', exc); return None


def _3d_3ds(path: str) -> Optional[str]:
    try:
        with open(path,'rb') as f:
            data = f.read()
        if data[:2] != b'\x4D\x4D': return None
        mesh_count = lights = cameras = 0
        i = 0
        while i < len(data) - 6:
            chunk_id  = struct.unpack_from('<H', data, i)[0]
            chunk_len = struct.unpack_from('<I', data, i+2)[0]
            if chunk_id == 0x4000: mesh_count += 1
            if chunk_id == 0x4600: lights     += 1
            if chunk_id == 0x4700: cameras    += 1
            if chunk_len < 6: break
            i += chunk_len
        parts = [f"3DS, {mesh_count} objects"]
        if lights:  parts.append(f"{lights} lights")
        if cameras: parts.append(f"{cameras} cameras")
        return ', '.join(parts)
    except Exception as exc:
        _log('_3d_3ds', exc); return None


def _3d_step(path: str) -> Optional[str]:
    try:
        with open(path,'r',errors='replace') as f:
            content = f.read()
        shells = content.count('CLOSED_SHELL')
        breps  = content.count('ADVANCED_BREP_SHAPE_REPRESENTATION')
        faces  = content.count('ADVANCED_FACE')
        parts  = ["STEP"]
        if breps:  parts.append(f"{breps} bodies")
        if shells: parts.append(f"{shells} shells")
        if faces:  parts.append(f"{faces} faces")
        return ', '.join(parts) if len(parts) > 1 else "STEP file"
    except Exception as exc:
        _log('_3d_step', exc); return None


def _3d_off(path: str) -> Optional[str]:
    try:
        with open(path,'r',errors='replace') as f:
            lines = [l.strip() for l in f if l.strip() and not l.startswith('#')]
        if not lines or lines[0] not in ('OFF','COFF','NOFF','CNOFF'): return None
        parts_of_line = lines[1].split()
        verts, faces = int(parts_of_line[0]), int(parts_of_line[1])
        return f"OFF: {verts:,} verts, {faces:,} faces"
    except Exception as exc:
        _log('_3d_off', exc); return None


def _3d_amf(path: str) -> Optional[str]:
    try:
        root = ET.parse(path).getroot()
        ns   = 'http://www.astm.org/cdr/schema/amf'
        objs = len(root.findall(f'{{{ns}}}object') or root.findall('.//object'))
        msh  = len(root.findall(f'{{{ns}}}mesh')   or root.findall('.//mesh'))
        return f"AMF: {objs} objects, {msh} meshes"
    except Exception as exc:
        _log('_3d_amf', exc); return None


def _3d_x3d(path: str) -> Optional[str]:
    try:
        root = ET.parse(path).getroot()
        shapes = len(list(root.iter('Shape')))
        tforms = len(list(root.iter('Transform')))
        return f"X3D: {shapes} shapes, {tforms} transforms"
    except Exception as exc:
        _log('_3d_x3d', exc); return None
